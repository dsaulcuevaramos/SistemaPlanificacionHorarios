from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func, delete
from sqlalchemy.orm import joinedload

from importlib import import_module

from app.core.database import get_db
# Modelos
from app.models.horario import Horario
from app.models.sesion import Sesion
from app.models.grupo import Grupo
from app.models.aula import Aula
from app.models.curso import Curso
from app.models.bloque_horario import BloqueHorario
from app.models.restriccion import Restriccion
from app.models.periodo_academico import PeriodoAcademico
from app.models.curso_aperturado import CursoAperturado
# Schemas
# Asegúrate de importar SesionFullResponse donde lo hayas definido
from app.schemas.sesion_completa import SesionFullResponse
from app.schemas.horario import HorarioCreate, HorarioResponse
from app.schemas.bloque_horario import BloqueHorarioResponse, BloqueMasivoCreate
from app.schemas.sesion import SesionResponse

# CRUDs
from app.crud.crud_sesion import sesion as crud_sesion
from app.crud.crud_bloque import bloque as crud_bloque

from fastapi.responses import StreamingResponse
import pandas as pd
import traceback
from io import BytesIO
from sqlalchemy.exc import IntegrityError
from sqlalchemy.dialects.postgresql import insert as pg_insert
from app.core.motor_horario import GeneradorHorario # Tu motor lógico

router = APIRouter()

# =====================================================================
#  1. LÓGICA DE VALIDACIÓN (EL ÁRBITRO)
# =====================================================================

async def validar_reglas_asignacion(
    db: AsyncSession, 
    id_sesion: int, 
    id_bloque: int, 
    id_periodo: int,
    id_aula: Optional[int] = None
) -> Optional[str]:
    """
    Verifica TODAS las reglas antes de permitir guardar un horario.
    Retorna None si es válido, o un string con el error.
    """
    
    # A. Datos de la Sesión
    stmt_sesion = select(Sesion).options(joinedload(Sesion.grupo)).where(Sesion.id == id_sesion)
    result_sesion = await db.execute(stmt_sesion)
    sesion_obj = result_sesion.scalar_one_or_none()
    
    if not sesion_obj: return "La sesión no existe."

    # B. REGLA 1: ANTI-BUCLE (Ya está completa?)
    stmt_count = select(func.count(Horario.id)).where(Horario.id_sesion == id_sesion,Horario.estado == 1)
    bloques_asignados = (await db.execute(stmt_count)).scalar() or 0
    
    if bloques_asignados >= sesion_obj.duracion_horas:
        return f"La sesión ya está completa ({bloques_asignados}/{sesion_obj.duracion_horas} horas). No puedes agregar más."

    # C. REGLA 2: CRUCES (Choques)
    id_grupo = sesion_obj.id_grupo
    id_docente = sesion_obj.grupo.id_docente

    conditions = [
        Horario.id_periodo == id_periodo,
        Horario.id_bloque == id_bloque,
        Horario.estado == 1
    ]
    
    conflict_or = [Sesion.id_grupo == id_grupo] # El grupo ya tiene clase
    if id_docente:
        conflict_or.append(and_(Grupo.id_docente == id_docente, Grupo.id_docente.isnot(None)))
    if id_aula:
        conflict_or.append(Horario.id_aula == id_aula)

    stmt_cruce = (
        select(Horario)
        .join(Sesion).join(Grupo)
        .where(and_(*conditions, or_(*conflict_or)))
        .options(joinedload(Horario.sesion).joinedload(Sesion.grupo).joinedload(Grupo.docente))
    )
    cruce = (await db.execute(stmt_cruce)).scalar_one_or_none()

    if cruce:
        detalles = []
        if cruce.sesion.id_grupo == id_grupo: detalles.append(f"El Grupo ya tiene clase ({cruce.sesion.tipo_sesion})")
        if id_docente and cruce.sesion.grupo.id_docente == id_docente: detalles.append(f"El Docente {cruce.sesion.grupo.docente.apellido} ya tiene clase")
        return "CRUCE: " + " | ".join(detalles)

    # D. REGLA 3: RESTRICCIONES (Días bloqueados)
    if id_docente:
        bloque_obj = await db.get(BloqueHorario, id_bloque)
        stmt_rest = select(Restriccion).where(
            Restriccion.entidad_referencia == 'DOCENTE',
            Restriccion.id_entidad == id_docente,
            Restriccion.tipo == 'BLOQUEO_DIA',
            Restriccion.estado == 1
        )
        restricciones = (await db.execute(stmt_rest)).scalars().all()
        for r in restricciones:
            if r.regla_json.get('dia') == bloque_obj.dia_semana:
                return f"El docente tiene restricción el {bloque_obj.dia_semana}."

    return None

# =====================================================================
#  2. ENDPOINTS DE DATOS (GET) - LO QUE TE FALTABA
# =====================================================================

@router.get("/sesiones/pendientes/{id_periodo}")
async def get_sesiones_pendientes(id_periodo: int, db: AsyncSession = Depends(get_db)):
    # Buscamos sesiones activas que NO estén en la tabla horario
    stmt = (
        select(Sesion)
        .join(Grupo).join(CursoAperturado).join(Curso)
        .outerjoin(Horario) # Left join
        .where(
            CursoAperturado.id_periodo == id_periodo,
            Sesion.estado == 1,
            Horario.id == None # Solo las que no tienen horario
        )
        .options(
            # ¡CRUCIAL! Cargar toda la jerarquía para el filtro del frontend
            joinedload(Sesion.grupo).joinedload(Grupo.curso_aperturado).joinedload(CursoAperturado.curso),
            joinedload(Sesion.grupo).joinedload(Grupo.docente),
            joinedload(Sesion.grupo).joinedload(Grupo.turno)
        )
    )
    result = await db.execute(stmt)
    return result.unique().scalars().all()

@router.get("/bloques/turno/{id_turno}", response_model=List[BloqueHorarioResponse])
async def read_bloques(id_turno: int, db: AsyncSession = Depends(get_db)):
    """Trae los bloques (horas) de un turno específico."""
    return await crud_bloque.get_by_turno(db, id_turno=id_turno)

@router.get("/periodo/{id_periodo}", response_model=List[HorarioResponse])
async def get_horario_periodo(id_periodo: int, db: AsyncSession = Depends(get_db)):
    """Trae todo el horario ya asignado (Fichas en la grilla)."""
    stmt = (
        select(Horario)
        .where(Horario.id_periodo == id_periodo, Horario.estado == 1)
        .options(
            joinedload(Horario.sesion).joinedload(Sesion.grupo).joinedload(Grupo.curso_aperturado).joinedload(import_module('app.models.curso_aperturado').CursoAperturado.curso),
            joinedload(Horario.sesion).joinedload(Sesion.grupo).joinedload(Grupo.docente),
            joinedload(Horario.bloque_horario),
            joinedload(Horario.aula)
        )
    )
    return (await db.execute(stmt)).scalars().all()

@router.get("/periodo/{id_periodo}/ciclos")
async def get_ciclos_en_periodo(id_periodo: int, db: AsyncSession = Depends(get_db)):
    """
    Retorna [1, 3, 5...] o [2, 4, 6...] según la paridad del periodo.
    """
    periodo = await db.get(PeriodoAcademico, id_periodo)
    if not periodo: return []
    
    codigo = str(periodo.codigo).upper().strip()
    es_par = codigo.endswith('II') or codigo.endswith('2') or 'PAR' in codigo
    ciclos = [2, 4, 6, 8, 10] if es_par else [1, 3, 5, 7, 9]
    
    return [{"id": c, "nombre": f"Ciclo {c}"} for c in ciclos]



@router.get("/sesiones/pendientes/{id_periodo}")
async def read_sesiones_pendientes(
    id_periodo: int,
    db: AsyncSession = Depends(get_db)
):
    stmt = (
        select(Sesion)
        .join(Grupo).join(CursoAperturado).join(Curso)
        .outerjoin(Horario) # Left join para ver si tiene horario
        .where(
            CursoAperturado.id_periodo == id_periodo,
            Horario.id == None, # Solo las que NO tienen horario
            Sesion.estado == 1
        )
        .options(
            # CARGA PROFUNDA OBLIGATORIA
            joinedload(Sesion.grupo).joinedload(Grupo.curso_aperturado).joinedload(CursoAperturado.curso),
            joinedload(Sesion.grupo).joinedload(Grupo.docente),
            joinedload(Sesion.grupo).joinedload(Grupo.turno) 
        )
    )
    result = await db.execute(stmt)
    # .unique() es vital cuando haces joins profundos que podrían duplicar filas en la query raw
    return result.unique().scalars().all()


# ==========================================
# 3. GUARDAR ASIGNACIÓN MANUAL (Validaciones)
# ==========================================
@router.post("/guardar-asignacion")
async def guardar_asignacion_manual(
    data: HorarioCreate,
    db: AsyncSession = Depends(get_db)
):
    stmt_sesion = (
        select(Sesion)
        .join(Grupo).join(CursoAperturado).join(Curso)
        .options(
            joinedload(Sesion.grupo).joinedload(Grupo.curso_aperturado).joinedload(CursoAperturado.curso),
            joinedload(Sesion.grupo).joinedload(Grupo.docente)
        )
        .where(Sesion.id == data.id_sesion)
    )
    sesion_actual = (await db.execute(stmt_sesion)).scalar()
    if not sesion_actual: raise HTTPException(404, "Sesión no encontrada")
    
    mi_ciclo = sesion_actual.grupo.curso_aperturado.curso.ciclo
    mi_grupo_nombre = sesion_actual.grupo.nombre
    duracion = sesion_actual.duracion_horas

    bloque_inicio = await db.get(BloqueHorario, data.id_bloque)
    if not bloque_inicio: raise HTTPException(404, "Bloque no encontrado")

    # Validar espacio en el turno
    stmt_bloques = (
        select(BloqueHorario)
        .where(
            BloqueHorario.id_turno == bloque_inicio.id_turno,
            BloqueHorario.dia_semana == bloque_inicio.dia_semana,
            BloqueHorario.orden >= bloque_inicio.orden
        )
        .order_by(BloqueHorario.orden)
        .limit(duracion)
    )
    bloques_validos = (await db.execute(stmt_bloques)).scalars().all()

    if len(bloques_validos) < duracion:
        raise HTTPException(400, f"No hay espacio suficiente. La clase dura {duracion} horas.")

    ids_bloques_necesarios = [b.id for b in bloques_validos]

    # Validar Docente
    if sesion_actual.grupo.id_docente:
        stmt_cruce = select(Horario).join(Sesion).join(Grupo).where(
            Horario.id_periodo == data.id_periodo,
            Horario.id_bloque.in_(ids_bloques_necesarios),
            Horario.estado == 1,
            Grupo.id_docente == sesion_actual.grupo.id_docente,
            Sesion.id != sesion_actual.id 
        )
        if (await db.execute(stmt_cruce)).first():
            raise HTTPException(400, "El Docente ya tiene clase asignada (cruce).")

    # Insertar o Actualizar
    stmt_existentes = select(Horario).where(
        Horario.id_periodo == data.id_periodo,
        Horario.ciclo == mi_ciclo,
        Horario.grupo == mi_grupo_nombre,
        Horario.id_bloque.in_(ids_bloques_necesarios)
    )
    casillas_existentes = (await db.execute(stmt_existentes)).scalars().all()
    mapa_existentes = {h.id_bloque: h for h in casillas_existentes}

    operaciones_update = []
    operaciones_insert = []

    for id_bloque_target in ids_bloques_necesarios:
        casilla = mapa_existentes.get(id_bloque_target)
        if casilla:
            if casilla.id_sesion is not None and casilla.id_sesion != data.id_sesion:
                 raise HTTPException(400, "Este espacio ya está ocupado por otro curso del grupo.")
            casilla.id_sesion = data.id_sesion
            casilla.id_aula = data.id_aula
            casilla.estado = 1
            operaciones_update.append(casilla)
        else:
            nuevo = Horario(
                id_sesion=data.id_sesion, id_bloque=id_bloque_target, id_periodo=data.id_periodo,
                id_aula=data.id_aula, ciclo=mi_ciclo, grupo=mi_grupo_nombre, estado=1
            )
            operaciones_insert.append(nuevo)

    try:
        if operaciones_insert: db.add_all(operaciones_insert)
        for up in operaciones_update: db.add(up)
        await db.commit()
        return {"message": "Asignado correctamente"}
    except IntegrityError:
        await db.rollback()
        raise HTTPException(400, "Error: Cruce detectado.")
    except Exception as e:
        await db.rollback()
        raise HTTPException(500, str(e))



@router.delete("/{id_horario}")
async def eliminar_asignacion(id_horario: int, db: AsyncSession = Depends(get_db)):
    # 1. Buscamos el horario
    horario = await db.get(Horario, id_horario)
    if not horario:
        raise HTTPException(404, "Horario no encontrado")

    # 2. ESTRATEGIA: "Limpiar" en lugar de "Borrar"
    # Si borramos la fila, perdemos el espacio en la grilla (el esqueleto).
    # Lo que queremos es quitar la sesión (id_sesion = None) y liberar el espacio.
    
    horario.id_sesion = None
    horario.id_aula = None
    horario.estado = 1 # Disponible / Vacío
    
    # IMPORTANTE: No tocamos 'ciclo', 'grupo' ni 'id_bloque' para que la celda
    # siga existiendo en la base de datos y el Frontend pueda dibujarla vacía.
    
    await db.commit()
    return {"message": "Asignación retirada correctamente (Espacio liberado)"}




# Nota: He resumido las funciones de exportación para no hacer el código gigante, 
# pero debes copiar el cuerpo de 'exportar_horario_excel' y 'exportar_plantilla_horarios' 
# de tu archivo gestion_horarios.py aquí mismo.



@router.post("/autogenerar-ciclo/{id_periodo}/{ciclo}")
async def autogenerar_por_ciclo(
    id_periodo: int,
    ciclo: int,
    db: AsyncSession = Depends(get_db)
):
    # A. BUSCAR PENDIENTES
    stmt_pendientes = (
        select(Sesion)
        .join(Grupo).join(CursoAperturado).join(Curso)
        .outerjoin(Horario) 
        .where(
            CursoAperturado.id_periodo == id_periodo,
            Curso.ciclo == ciclo,
            Horario.id == None, 
            Sesion.estado == 1
        )
        .options(
            joinedload(Sesion.grupo).joinedload(Grupo.curso_aperturado).joinedload(CursoAperturado.curso),
            joinedload(Sesion.grupo).joinedload(Grupo.docente),
            joinedload(Sesion.grupo).joinedload(Grupo.turno)
        )
    )
    sesiones_pendientes = (await db.execute(stmt_pendientes)).unique().scalars().all()
    
    if not sesiones_pendientes:
        return {"status": "info", "message": f"No hay sesiones pendientes para el Ciclo {ciclo}."}

    # B. BUSCAR OCUPADOS (BLINDAJE)
    stmt_ocupados = (
        select(Horario)
        .join(Sesion).join(Grupo)
        .where(Horario.id_periodo == id_periodo, Horario.estado == 1)
        .options(
            joinedload(Horario.bloque_horario), 
            joinedload(Horario.sesion).joinedload(Sesion.grupo)
        )
    )
    horarios_ocupados_objs = (await db.execute(stmt_ocupados)).scalars().all()
    
    lista_ocupados = []
    for h in horarios_ocupados_objs:
        if not h.bloque_horario: continue
        lista_ocupados.append({
            'dia': h.bloque_horario.dia_semana,
            'id_bloque': h.bloque_horario.orden,
            'id_docente': h.sesion.grupo.id_docente, 
            'grupo_uid': h.sesion.grupo.id 
        })

    # C. MOTOR
    data_sesiones = []
    for s in sesiones_pendientes:
        data_sesiones.append({
            "ID_SESION": s.id,
            "ID_DOCENTE": s.grupo.id_docente,
            "GRUPO": s.grupo.nombre,
            "GRUPO_UID": s.grupo.id,
            "TURNO_GRUPO": s.grupo.turno.nombre,
            "DURACION_HORAS": s.duracion_horas,
            "CURSO": s.grupo.curso_aperturado.curso.nombre,
            "CICLO": s.grupo.curso_aperturado.curso.ciclo, 
            "DIA": None, "BLOQUE_ORDEN": None
        })
    df_sesiones = pd.DataFrame(data_sesiones)
    
    stmt_bloques = select(BloqueHorario).distinct(BloqueHorario.orden)
    bloques_result = (await db.execute(stmt_bloques)).scalars().all()
    df_bloques = pd.DataFrame([{'orden': b.orden} for b in bloques_result])

    motor = GeneradorHorario(df_sesiones, df_bloques, horarios_ocupados=lista_ocupados)
    df_resultado, fallos = motor.ejecutar()

    # D. GUARDADO (UPSERT)
    all_bloques = (await db.execute(select(BloqueHorario))).scalars().all()
    mapa_bloques = {}
    for b in all_bloques:
        mapa_bloques[(b.dia_semana, b.orden, b.id_turno)] = b.id

    mapa_sesion_turno = {s.id: s.grupo.id_turno for s in sesiones_pendientes}
    lista_para_upsert = []

    for _, row in df_resultado.iterrows():
        if row['DIA'] and row['BLOQUE_ORDEN']:
            duracion = int(row['DURACION_HORAS'])
            orden_inicio = int(row['BLOQUE_ORDEN'])
            dia = row['DIA']
            id_sesion = int(row['ID_SESION'])
            nombre_grupo = str(row['GRUPO'])
            ciclo_actual = int(row['CICLO'])
            id_turno_grupo = mapa_sesion_turno.get(id_sesion)

            for i in range(duracion):
                orden_actual = orden_inicio + i
                id_bloque = mapa_bloques.get((dia, orden_actual, id_turno_grupo))
                
                if id_bloque:
                    lista_para_upsert.append({
                        "id_periodo": id_periodo,
                        "id_bloque": id_bloque,
                        "ciclo": ciclo_actual,
                        "grupo": nombre_grupo,
                        "id_sesion": id_sesion,
                        "estado": 1,
                        "id_aula": None 
                    })
    
    if lista_para_upsert:
        stmt = pg_insert(Horario).values(lista_para_upsert)
        stmt = stmt.on_conflict_do_update(
            constraint='uq_horario_casilla', 
            set_={ "id_sesion": stmt.excluded.id_sesion, "estado": 1 }
        )
        try:
            await db.execute(stmt)
            await db.commit()
        except Exception as e:
            await db.rollback()
            raise HTTPException(500, f"Error en BD: {str(e)}")

    return {"status": "success", "generados": len(lista_para_upsert), "fallos": len(fallos)}



def safe_to_roman(n):
    if not isinstance(n, int) or n < 1: return "0"
    mapa = ["", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
    return mapa[n] if n <= 10 else str(n)



@router.get("/exportar-excel/{id_periodo}")
async def exportar_horario_excel(
    id_periodo: int,
    db: AsyncSession = Depends(get_db)
):
    try:
        periodo_obj = await db.get(PeriodoAcademico, id_periodo)
        nombre_semestre = periodo_obj.nombre if periodo_obj else "202X"

        stmt_grupos = (
            select(Grupo)
            .join(CursoAperturado).join(Curso)
            .where(CursoAperturado.id_periodo == id_periodo)
            .options(joinedload(Grupo.curso_aperturado).joinedload(CursoAperturado.curso), joinedload(Grupo.docente))
            .order_by(Curso.ciclo, Grupo.nombre)
        )
        grupos = (await db.execute(stmt_grupos)).scalars().all()

        if not grupos:
            df_vacio = pd.DataFrame(["No hay grupos registrados"])
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_vacio.to_excel(writer, sheet_name='Info')
            output.seek(0)
            return StreamingResponse(output, headers={'Content-Disposition': 'attachment; filename="Vacio.xlsx"'})

        stmt_horarios = (
            select(Horario)
            .join(Sesion).join(Grupo).join(BloqueHorario).outerjoin(Aula)
            .where(Horario.id_periodo == id_periodo, Horario.estado == 1)
            .options(
                joinedload(Horario.bloque_horario), joinedload(Horario.aula),
                joinedload(Horario.sesion).joinedload(Sesion.grupo).joinedload(Grupo.docente),
                joinedload(Horario.sesion).joinedload(Sesion.grupo).joinedload(Grupo.curso_aperturado).joinedload(CursoAperturado.curso)
            )
        )
        asignaciones = (await db.execute(stmt_horarios)).scalars().all()

        mapa_contenido = {}
        for h in asignaciones:
            if not h.bloque_horario: continue
            key = (h.sesion.id_grupo, h.bloque_horario.dia_semana, h.bloque_horario.orden)
            curso = h.sesion.grupo.curso_aperturado.curso.nombre
            docente = h.sesion.grupo.docente.apellido if h.sesion.grupo.docente else "(VACANTE)"
            aula = f"[{h.aula.codigo}]" if h.aula else ""
            mapa_contenido[key] = f"{curso}\n{docente} {aula}"

        stmt_bloques = select(BloqueHorario).distinct(BloqueHorario.orden).order_by(BloqueHorario.orden)
        bloques_unicos = (await db.execute(stmt_bloques)).scalars().all()
        dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]

        data_excel = []
        for g in grupos:
            ciclo_num = g.curso_aperturado.curso.ciclo
            lbl_ciclo = f"CICLO {safe_to_roman(ciclo_num)}"
            lbl_grupo = f"GRUPO {g.nombre}"
            
            mis_bloques = [b for b in bloques_unicos if b.id_turno == g.id_turno] or bloques_unicos

            for dia in dias_semana:
                for b in mis_bloques:
                    hora_fmt = f"{b.hora_inicio.strftime('%H:%M')} - {b.hora_fin.strftime('%H:%M')}"
                    val = mapa_contenido.get((g.id, dia, b.orden), "")
                    data_excel.append({
                        "CicloSort": ciclo_num, "Ciclo": lbl_ciclo, "Grupo": lbl_grupo,
                        "Dia": dia, "Hora": hora_fmt, "Orden": b.orden, "Contenido": val
                    })

        df = pd.DataFrame(data_excel)
        df['Dia'] = pd.Categorical(df['Dia'], categories=dias_semana, ordered=True)
        df.sort_values(by=['CicloSort', 'Grupo', 'Dia', 'Orden'], inplace=True)

        pivot_df = df.pivot_table(
            index=["Dia", "Hora"], columns=["Ciclo", "Grupo"], values="Contenido", 
            aggfunc=lambda x: ' / '.join([str(v) for v in x if v]), sort=False
        )

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pivot_df.to_excel(writer, sheet_name='Horario General', startrow=6)
            ws = writer.sheets['Horario General']
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            # Títulos
            titulos = [
                "UNIVERSIDAD NACIONAL DE UCAYALI",
                "FACULTAD DE CIENCIAS ECONOMICAS ADMINISTRATIVAS Y CONTABLES",
                "ESCUELA PROFESIONAL DE CIENCIAS ADMINISTRATIVAS",
                f"SEMESTRE ACADÉMICO {nombre_semestre} - HORARIO GENERAL"
            ]
            total_cols = ws.max_column
            for i, txt in enumerate(titulos, start=1):
                cell = ws.cell(row=i, column=1)
                cell.value = txt
                cell.font = Font(name='Arial', size=12, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=total_cols)

            # Estilos
            thin = Side(style='thin', color="000000")
            border_full = Border(left=thin, right=thin, top=thin, bottom=thin)
            fill_header_ciclo = PatternFill("solid", fgColor="366092")
            fill_header_grupo = PatternFill("solid", fgColor="B8CCE4")
            font_white = Font(color="FFFFFF", bold=True)
            font_black = Font(color="000000", bold=True)

            for row in ws.iter_rows(min_row=7):
                for cell in row:
                    cell.border = border_full
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    if cell.row == 7 and cell.value:
                        cell.fill = fill_header_ciclo
                        cell.font = font_white
                    elif cell.row == 8 and cell.column > 2:
                        cell.fill = fill_header_grupo
                        cell.font = font_black
                    if cell.column <= 2:
                        cell.font = font_black
                        cell.fill = PatternFill("solid", fgColor="F2F2F2")

            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 15
            for col in range(3, total_cols + 1):
                col_letter = ws.cell(row=8, column=col).column_letter
                ws.column_dimensions[col_letter].width = 25

        output.seek(0)
        return StreamingResponse(
            output, 
            headers={'Content-Disposition': f'attachment; filename="Horario_General_{id_periodo}.xlsx"'},
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print("❌ ERROR EXPORTAR EXCEL:")
        traceback.print_exc()
        raise HTTPException(500, f"Error interno: {str(e)}")

# ==========================================
# 2. ENDPOINTS DE GESTIÓN (Bloques y Sesiones)
# ==========================================

@router.get("/bloques/turno/{id_turno}", response_model = List[BloqueHorarioResponse])
async def read_bloques(id_turno: int, db: AsyncSession = Depends(get_db)):
    return await crud_bloque.get_by_turno(db, id_turno=id_turno)

@router.post("/bloques/masivo")
async def create_bloques_masivos(payload: BloqueMasivoCreate, db: AsyncSession = Depends(get_db)):
    return await crud_bloque.create_bulk(db, payload.id_turno, payload.dias, payload.intervalos)

@router.delete("/bloques/turno/{id_turno}")
async def clear_blocks_by_turno(id_turno: int, db: AsyncSession = Depends(get_db)):
    await crud_bloque.remove_by_turno(db, id_turno=id_turno)
    return {"detail": "Rejilla limpiada"}

@router.get("/sesiones/grupo/{id_grupo}", response_model=List[SesionResponse])
async def get_sesiones_by_grupo(id_grupo: int, db: AsyncSession = Depends(get_db)):
    stmt = select(Sesion).where(Sesion.id_grupo == id_grupo)
    return (await db.execute(stmt)).scalars().all()

@router.post("/sesiones/generar-automatico/{id_grupo}")
async def auto_generar_sesiones(id_grupo: int, db: AsyncSession = Depends(get_db)):
    return await crud_sesion.generar_sesiones_por_grupo(db, id_grupo=id_grupo)
