import pandas as pd
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_
from typing import Optional
from typing import List
from sqlalchemy.orm import joinedload
from app.core.database import get_db
from app.api.deps import get_current_user

from app.models.usuario import Usuario
from app.models.horario import Horario
from app.schemas.horario import HorarioCreate
from app.schemas.bloque_horario import BloqueHorarioCreate, BloqueHorarioResponse, BloqueMasivoCreate
from app.models.sesion import Sesion
from app.schemas.sesion import SesionResponse
from app.models.grupo import Grupo  
from app.models.curso_aperturado import CursoAperturado
from app.crud.crud_sesion import sesion as crud_sesion
from app.crud.crud_bloque import bloque as crud_bloque # Asegúrate de tener esto importado

from app.models.periodo_academico import PeriodoAcademico
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from app.crud.crud_sesion import sesion as crud_sesion


from app.core.motor_horario import GeneradorHorario
from app.models.bloque_horario import BloqueHorario
from app.models.aula import Aula

router = APIRouter()

# --- FUNCIÓN DE VALIDACIÓN COMPLETA ---
async def verificar_colisiones_universidad(
    db: AsyncSession, 
    id_sesion: int, 
    id_bloque: int, 
    id_periodo: int,
    id_aula: Optional[int] = None # Aula es opcional ahora
) -> Optional[str]:
    """
    Valida las reglas de negocio universitarias:
    1. GRUPO: Los alumnos no pueden tener 2 clases al mismo tiempo. (Crítico)
    2. DOCENTE: Si el grupo tiene docente, este no puede estar en otro lado.
    3. AULA: Si se asigna aula, no puede estar ocupada.
    """
    
    # 1. Obtener datos completos de la sesión que queremos programar
    sesion_actual = await crud_sesion.get_con_detalles(db, id_sesion)
    if not sesion_actual:
        return f"Sesión {id_sesion} no encontrada."

    grupo = sesion_actual.grupo
    id_docente = grupo.id_docente
    id_grupo = grupo.id

    # 2. Construir la consulta de conflictos en la tabla Horario
    # Buscamos registros en el MISMO periodo y MISMO bloque
    conditions = [
        Horario.id_periodo == id_periodo,
        Horario.id_bloque == id_bloque,
        Horario.estado == 1
    ]

    # Sub-condiciones de choque (OR)
    conflict_conditions = [Sesion.id_grupo == id_grupo]

    # A) Choque de Grupo (Los alumnos ya están ocupados)
    conflict_conditions.append(Sesion.id_grupo == id_grupo)

    # B) Choque de Docente (Solo si el docente existe)
    if id_docente is not None:
        conflict_conditions.append(
            and_(Grupo.id_docente == id_docente, Grupo.id_docente.isnot(None))
        )

    # C) Choque de Aula (Solo si estamos intentando asignar una aula)
    if id_aula is not None:
        conflict_conditions.append(Horario.id_aula == id_aula)

    # Unir todo
    stmt = (
        select(Horario)
        .join(Sesion).join(Grupo)
        .where(
            and_(
                Horario.id_periodo == id_periodo,
                Horario.id_bloque == id_bloque,
                Horario.estado == 1,
                or_(*conflict_conditions) # Se activa cualquier regla que coincida
            )
        )
    )

    result = await db.execute(stmt)
    choque = result.scalar_one_or_none()

    # 3. Diagnóstico del error (si hubo choque)
    if choque:
        motivo = []
        if choque.sesion.id_grupo == id_grupo:
            motivo.append(f"El Grupo '{grupo.nombre}' ya tiene clase ({choque.sesion.tipo_sesion}).")
        
        if id_docente and choque.sesion.grupo.id_docente == id_docente:
            nombre_docente = grupo.docente.nombre if grupo.docente else "Docente"
            motivo.append(f"El docente {nombre_docente} ya dicta en el grupo '{choque.sesion.grupo.nombre}'.")
        
        if id_aula and choque.id_aula == id_aula:
            motivo.append(f"El aula ID {id_aula} ya está ocupada.")

        return " | ".join(motivo)

    return None # Sin conflictos


# ==========================================
# 2. ENDPOINTS DE GESTIÓN (Bloques y Sesiones)
# ==========================================

@router.get("/bloques/turno/{id_turno}", response_model = List[BloqueHorarioResponse])
async def read_bloques(id_turno: int, db: AsyncSession = Depends(get_db)):
    return await crud_bloque.get_by_turno(db, id_turno=id_turno)

@router.post("/bloques/masivo")
async def create_bloques_masivos(payload: BloqueMasivoCreate, db: AsyncSession = Depends(get_db)):
    return await crud_bloque.create_bulk(db, payload.id_turno, payload.dias, payload.intervalos)

@router.delete("/bloques/turno/{id_turno}")
async def clear_blocks_by_turno(id_turno: int, db: AsyncSession = Depends(get_db)):
    await crud_bloque.remove_by_turno(db, id_turno=id_turno)
    return {"detail": "Rejilla limpiada"}

@router.get("/sesiones/grupo/{id_grupo}", response_model=List[SesionResponse])
async def get_sesiones_by_grupo(id_grupo: int, db: AsyncSession = Depends(get_db)):
    stmt = select(Sesion).where(Sesion.id_grupo == id_grupo)
    return (await db.execute(stmt)).scalars().all()

@router.post("/sesiones/generar-automatico/{id_grupo}")
async def auto_generar_sesiones(id_grupo: int, db: AsyncSession = Depends(get_db)):
    return await crud_sesion.generar_sesiones_por_grupo(db, id_grupo=id_grupo)



# ==========================================
# 3. ENDPOINTS DE ASIGNACIÓN (Horario)
# ==========================================

@router.post("/horarios/asignar")
async def asignar_horario(
    payload: HorarioCreate, 
    db: AsyncSession = Depends(get_db)
):
    error = await verificar_colisiones_universidad(
        db, payload.id_sesion, payload.id_bloque, payload.id_periodo, payload.id_aula
    )
    if error:
        raise HTTPException(status_code=400, detail=error)

    nuevo = Horario(**payload.dict())
    db.add(nuevo)
    await db.commit()
    return nuevo


# --- ENDPOINT 1: EXPORTAR PLANTILLA CSV ---
@router.get("/exportar-plantilla/{id_periodo}")
async def exportar_plantilla_horarios(
    id_periodo: int,
    db: AsyncSession = Depends(get_db)
):
    """
    Genera el CSV Base (Borrador) para el periodo.
    Columnas de DATOS (Info del curso/grupo) + Columnas de TRABAJO (Dia, Bloque, Aula).
    """
    # 1. Obtener nombre del periodo para el archivo
    periodo = await db.get(PeriodoAcademico, id_periodo)
    nombre_archivo = f"Horario_Borrador_{periodo.codigo}.csv" if periodo else "Horario_Borrador.csv"

    # 2. Obtener sesiones pendientes (o todas)
    sesiones = await crud_sesion.get_pendientes_export(db, id_periodo)
    
    data = []
    for s in sesiones:
        # Recuperar nombres para referencia visual
        nombre_docente = s.grupo.docente.apellido + " " + s.grupo.docente.nombre if s.grupo.docente else "VACANTE"
        
        data.append({
            # --- COLUMNAS DE REFERENCIA (NO EDITAR EN EXCEL) ---
            "ID_SESION": s.id,  # CLAVE ÚNICA PARA EL IMPORTADOR
            "CODIGO_CURSO": s.grupo.curso_aperturado.curso.codigo,
            "CURSO": s.grupo.curso_aperturado.curso.nombre,
            "GRUPO": s.grupo.nombre,
            "TURNO_GRUPO": s.grupo.turno.nombre if s.grupo.turno else "Sin Turno",
            "TIPO": s.tipo_sesion,
            "DURACION_HORAS": s.duracion_horas,
            "DOCENTE": nombre_docente,
            
            # --- COLUMNAS DE TRABAJO (AQUÍ RELLENA EL ALGORITMO O EL USUARIO) ---
            "DIA": "",           # Ej: Lunes, Martes...
            "BLOQUE_ORDEN": "",  # Ej: 1, 2, 3 (Según tu configuración de bloques)
            "ID_AULA": ""        # Opcional: ID del aula física
        })

    # 3. Generar CSV
    df = pd.DataFrame(data)
    stream = io.StringIO()
    # index=False para no guardar el número de fila de Pandas
    df.to_csv(stream, index=False, encoding='utf-8-sig') # utf-8-sig para que Excel abra bien las tildes
    
    response = StreamingResponse(iter([stream.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename={nombre_archivo}"
    
    return response

# --- ENDPOINT 2: IMPORTAR / VALIDAR CSV ---
@router.post("/importar-csv")
async def importar_csv(
    id_periodo: int,
    file: UploadFile = File(...),
    confirmar_guardado: bool = False,
    db: AsyncSession = Depends(get_db)
):
    # Validaciones básicas de archivo...
    content = await file.read()
    df = pd.read_csv(io.BytesIO(content))
    
    # Mapa de optimización
    mapa_bloques = await crud_bloque.get_by_periodo_map(db, id_periodo)
    
    errores = []
    validos = []

    for idx, row in df.iterrows():
        try:
            # 1. Traducir CSV (Dia/Orden) -> ID Bloque
            dia, orden = row.get("DIA", "").capitalize(), int(row.get("BLOQUE_ORDEN", 0))
            id_bloque = mapa_bloques.get((dia, orden))
            
            if not id_bloque:
                errores.append({"fila": idx+2, "error": f"Bloque inválido: {dia} - {orden}"})
                continue
            
            # 2. Manejar Aula opcional (NaN -> None)
            id_aula = int(row["ID_AULA"]) if pd.notna(row.get("ID_AULA")) else None
            id_sesion = int(row["ID_SESION"])

            # 3. Validar con lógica Universitaria
            msg = await verificar_colisiones_universidad(db, id_sesion, id_bloque, id_periodo, id_aula)
            
            if msg:
                errores.append({"fila": idx+2, "error": msg})
            else:
                validos.append(Horario(
                    id_sesion=id_sesion, id_bloque=id_bloque, id_aula=id_aula, 
                    id_periodo=id_periodo, estado=1
                ))

        except Exception as e:
            errores.append({"fila": idx+2, "error": f"Error de formato: {str(e)}"})

    if errores:
        return {"status": "error", "detalles": errores}

    if confirmar_guardado:
        db.add_all(validos)
        await db.commit()
        return {"status": "success", "message": f"Guardados {len(validos)} horarios."}

    return {"status": "valid", "message": "Validación exitosa."}


# ==========================================
# 5. EXPORTACIÓN VISUAL (EXCEL)
# ==========================================

@router.get("/exportar-excel/grupo/{id_grupo}")
async def exportar_horario_excel(
    id_grupo: int,
    db: AsyncSession = Depends(get_db)
):
    """
    Genera un Excel visualmente organizado con el horario del grupo seleccionado.
    Ideal para compartir con alumnos y docentes.
    """
    # 1. Obtener Datos del Grupo
    stmt_grupo = select(Grupo).where(Grupo.id == id_grupo).options(joinedload(Grupo.turno))
    grupo = (await db.execute(stmt_grupo)).scalar_one_or_none()
    
    if not grupo:
        raise HTTPException(status_code=404, detail="Grupo no encontrado")

    # 2. Obtener Bloques y Horarios
    bloques = await crud_bloque.get_by_turno(db, id_turno=grupo.id_turno)
    
    stmt_horario = (
        select(Horario)
        .join(Sesion).join(Grupo)
        .where(Sesion.id_grupo == id_grupo, Horario.estado == 1)
        .options(
            joinedload(Horario.sesion).joinedload(Sesion.grupo).joinedload(Grupo.curso_aperturado).joinedload(CursoAperturado.curso),
            joinedload(Horario.sesion).joinedload(Sesion.grupo).joinedload(Grupo.docente),
            joinedload(Horario.aula)
        )
    )
    horarios = (await db.execute(stmt_horario)).scalars().all()

    # 3. Crear Excel con OpenPyXL
    wb = Workbook()
    ws = wb.active
    ws.title = f"Horario {grupo.nombre}"

    # --- ESTILOS ---
    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- ENCABEZADO DEL DOCUMENTO ---
    ws.merge_cells('A1:G1')
    ws['A1'] = f"HORARIO DE CLASES - {grupo.nombre} ({grupo.turno.nombre})"
    ws['A1'].font = Font(size=14, bold=True, color="1E293B")
    ws['A1'].alignment = center_align

    # --- CABECERA DE LA TABLA ---
    headers = ["HORA", "LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO"]
    ws.append(headers)
    
    # Aplicar estilos a la fila de cabecera (Fila 2)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # --- CONSTRUIR MATRIZ ---
    # Mapa rápido para encontrar clases: {(dia, id_bloque): horario_obj}
    mapa_horario = {(h.bloque.dia_semana, h.id_bloque): h for h in horarios}
    
    # Mapa de columnas
    dias_col = {"Lunes": 2, "Martes": 3, "Miércoles": 4, "Jueves": 5, "Viernes": 6, "Sábado": 7}

    row_idx = 3
    for b in bloques:
        # Columna 1: La Hora
        cell_hora = ws.cell(row=row_idx, column=1)
        cell_hora.value = f"{b.hora_inicio[:5]} - {b.hora_fin[:5]}"
        cell_hora.alignment = center_align
        cell_hora.font = Font(bold=True)
        cell_hora.border = thin_border

        # Columnas 2-7: Los Días
        for dia, col_idx in dias_col.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = center_align
            
            # Buscar si hay clase
            match = mapa_horario.get((dia, b.id))
            if match:
                curso = match.sesion.grupo.curso_aperturado.curso.nombre
                docente = match.sesion.grupo.docente.nombre if match.sesion.grupo.docente else "Sin Docente"
                aula = match.aula.codigo if match.aula else "Aula Pendiente"
                
                # Texto dentro de la celda (Salto de línea para orden)
                cell.value = f"{curso}\n({match.sesion.tipo_sesion})\n👨‍🏫 {docente}\n📍 {aula}"
                cell.fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid") # Fondo suave
            else:
                cell.value = "-"

        row_idx += 1

    # --- AJUSTAR ANCHO DE COLUMNAS ---
    ws.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 25

    # 4. Retornar Archivo
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"Horario_{grupo.nombre.replace(' ', '_')}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(stream, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/sesiones/pendientes/{id_periodo}", response_model=List[SesionResponse])
async def get_sesiones_pendientes(
    id_periodo: int,
    db: AsyncSession = Depends(get_db)
):
    """
    Retorna todas las sesiones del periodo.
    El frontend filtrará luego las que pertenecen al grupo seleccionado.
    """
    # Reutilizamos la lógica de exportación que ya trae las sesiones con sus relaciones
    return await crud_sesion.get_pendientes_export(db, id_periodo)


@router.post("/generar-automatico/{id_periodo}")
async def generar_horario_automatico(
    id_periodo: int,
    db: AsyncSession = Depends(get_db)
):
    """
    1. Lee todas las sesiones pendientes.
    2. Lee la configuración de bloques y aulas.
    3. Ejecuta el algoritmo de asignación.
    4. Retorna el CSV listo para revisión.
    """
    # A. Obtener Datos (Sesiones Pendientes)
    sesiones = await crud_sesion.get_pendientes_export(db, id_periodo)
    if not sesiones:
        raise HTTPException(400, "No hay sesiones pendientes para programar.")

    # Convertir a DataFrame (Formato que entiende el motor)
    data_sesiones = []
    for s in sesiones:
        nombre_docente = s.grupo.docente.apellido if s.grupo.docente else "VACANTE"
        data_sesiones.append({
            "ID_SESION": s.id,
            "CURSO": s.grupo.curso_aperturado.curso.nombre,
            "GRUPO": s.grupo.nombre,
            "TURNO_GRUPO": s.grupo.turno.nombre if s.grupo.turno else "General",
            "DOCENTE": nombre_docente,
            "DURACION_HORAS": s.duracion_horas,
            "DIA": "", "BLOQUE_ORDEN": "", "ID_AULA": ""
        })
    df_sesiones = pd.DataFrame(data_sesiones)

    # B. Obtener Configuración de Bloques (Para saber los límites)
    # Traemos todos los bloques únicos usados en este periodo (a través de los turnos)
    # Nota: Simplificamos trayendo todos los bloques posibles
    stmt_bloques = select(BloqueHorario.orden).distinct()
    bloques_result = await db.execute(stmt_bloques)
    df_bloques = pd.DataFrame(bloques_result.all(), columns=['orden'])

    # C. Obtener Aulas
    stmt_aulas = select(Aula.id, Aula.nombre, Aula.aforo)
    aulas_result = await db.execute(stmt_aulas)
    # Convertimos a lista de dicts para pandas
    aulas_data = [{"id": r.id, "nombre": r.nombre, "aforo": r.aforo} for r in aulas_result.all()]
    df_aulas = pd.DataFrame(aulas_data)

    # D. EJECUTAR MOTOR
    motor = GeneradorHorario(df_sesiones, df_bloques, df_aulas)
    df_resultado, fallos = motor.ejecutar()

    # E. Devolver CSV
    stream = io.StringIO()
    df_resultado.to_csv(stream, index=False, encoding='utf-8-sig')
    
    filename = f"Horario_Generado_Auto_{len(fallos)}_Fallos.csv"
    
    response = StreamingResponse(iter([stream.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    
    return response